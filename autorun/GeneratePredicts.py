import csv
import os

import openpyxl
import torch
from skimage import io
from tqdm import tqdm

from models.ModelsClasses import Covid
from models.ModelsLearner import CovidModels
from utils.Utils import create_path_if_not_exists

if __name__ == "__main__":
    modelsLearner = CovidModels()

    print("Founded models configurations:")

    path = "ModelsLearning.xlsx"

    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active

    row = sheet_obj.max_row
    configurations = dict()
    for i in range(2, row + 1):
        name = sheet_obj.cell(row=i, column=1)
        model_name = sheet_obj.cell(row=i, column=3)
        configurations.update({name.value: model_name.value})
        print(name.value)
    print()
    selectedModelName = input("Select: ")

    modelClass = modelsLearner.models[configurations[selectedModelName]].model()

    epoch = input("Enter epoch id: ")

    modelWeightsPath = f"../model_results/{selectedModelName}_variants/{selectedModelName}_{epoch}_epoch"

    train_images_path = "/Users/max/Downloads/data/test_images"
    predict_path = "../predicts"

    predict_filename = predict_path + f"/{selectedModelName}_{epoch}.csv"
    create_path_if_not_exists(predict_path)
    modelClass.load_state_dict(torch.load(modelWeightsPath))
    modelClass.to(torch.device('mps'))
    fields = ['Id', 'target_feature']
    with open(predict_filename, 'w') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fields)
        writer.writeheader()
        for idx in tqdm(range(6920)):
            img_name = os.path.join(train_images_path,
                                    "img_" + str(idx) + ".png")

            # Записыываем картинки в переменные
            image = torch.tensor(io.imread(img_name).reshape(1, 1, 256, 256)).to(torch.device('mps'))
            y = modelClass(image.to(torch.float32))
            writer.writerow({'Id': idx, 'target_feature': y.argmax(dim=1)[0].item()})
