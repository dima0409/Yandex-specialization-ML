import openpyxl
import torch

from dataTranforms.DataLoaders import DataLoadersGenerator
from models.ModelsLearner import CovidModels
from utils.Utils import fit
import getopt
import sys

if __name__ == "__main__":
    argv = sys.argv[1:]

    opts, args = getopt.getopt(argv, "s:d:",
                               ["dataset =", "device ="])

    csv_dir = ""
    original_dir = ""
    cropped_dir = ""
    device = ""
    for opt, arg in opts:
        if opt in ['-s', '--dataset']:
            csv_dir = arg + "train_answers.csv"
            original_dir = arg + "train_images"
            cropped_dir = arg + "train_lung_masks"
        elif opt in ['-d', '--device']:
            device = arg

    path = "autorun/ModelsLearning.xlsx"

    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active

    row = sheet_obj.max_row

    # set dataset paths
    dataLoadersGenerator = DataLoadersGenerator(csv_dir,
                                                original_dir,
                                                cropped_dir)
    models = CovidModels()
    device_obj = torch.device(device)  # set learning device (Apple Silicon - "mps"; GPU Cuda cores - "cuda"; CPU - "cpu")

    if row >= 2:
        for i in range(2, row + 1):
            version_name = sheet_obj.cell(row=i, column=1).value
            data_loader_name = sheet_obj.cell(row=i, column=2).value
            model_name = sheet_obj.cell(row=i, column=3).value
            status = sheet_obj.cell(row=i, column=4)

            if status.value in ("", None, "trainable"):
                print(f"Start learning {version_name}")

                status.value = "progress"
                wb_obj.save(path)

                dataloaders = dataLoadersGenerator.data_loaders[data_loader_name]
                model_info = models.models[model_name]
                model_obj = model_info.model()
                model_obj.to(device_obj)
                fit(model_obj, train_loader=dataloaders.train_dl, valid_loader=dataloaders.test_dl,
                    x_field_name=model_obj.x_field_name, y_field_name=model_obj.y_field_name,
                    loss_fn=model_info.loss_fn,
                    optimizer=model_info.optimizer(model_obj.parameters(), lr=model_info.lr),
                    num_epochs=model_info.epochs,
                    title=version_name, device=device_obj)

                wb_obj = openpyxl.load_workbook(path)
                wb_obj.active.cell(row=i, column=4).value = "done"

                print(f"Finished learning {version_name}")

    else:
        print("Empty xlsx")

    wb_obj.save(path)
